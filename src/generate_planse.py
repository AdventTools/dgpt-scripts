"""
generate_planse.py

Genereaza planse PPTX: 1 slide per (cabinet, medic) din data/medici.xlsx.

Input:  data/medici.xlsx       (coloane: cabinet, nume; randul 1 = headere)
Sursa:  templates/Planse.pptx  (1 slide cu 2 placeholdere: {{medic}} si {{cabinet}})
Output: output/Planse_cabinete.pptx

Pentru fiecare slide:
- run[0] = {{medic}}, dimensiune implicita 80pt (Impact). Auto-fit doar daca nu incape.
- <a:br> = line break (aceeasi dimensiune ca medicul)
- run[1] = {{cabinet}}, dimensiune implicita 115pt (Impact). Auto-fit doar daca nu incape.

Fontul Impact si dimensiunile maxime sunt respectate cand textul incape pe un rand.
Daca un nume/cabinet e prea lung, fontul scade pana la FONT_MIN_PT.
"""
import openpyxl
from pptx import Presentation

from pptx_utils import (
    DATA, TEMPLATES_DIR, OUTPUT_DIR,
    NS_A, NSMAP,
    IMPACT_PATH, auto_font_pt,
    clone_slide,
)

INPUT_XLSX  = DATA          / 'medici.xlsx'
TEMPLATE    = TEMPLATES_DIR / 'Planse.pptx'
OUTPUT_PPTX = OUTPUT_DIR    / 'Planse_cabinete.pptx'

# Latimea textbox-ului in puncte (cx=9963000 EMU, lIns=rIns=116050 EMU; 1pt=12700 EMU)
BOX_W_PT     = (9963000 - 116050 * 2) / 12700   # ≈ 766 pt
# Marja 4% ca textul Impact sa nu fie taiat la randare in PowerPoint
SAFE_BOX_PT  = BOX_W_PT * 0.96                  # ≈ 735 pt

# Dimensiuni maxime din template (se respecta daca textul incape pe un rand)
DOC_MAX_PT   = 80
CAB_MAX_PT   = 115
FONT_MIN_PT  = 28
IMPACT_RATIO = 0.52   # fallback pentru aproximare cand Pillow lipseste

# Placeholdere din template
PLACEHOLDER_MEDIC   = '{{medic}}'
PLACEHOLDER_CABINET = '{{cabinet}}'


def load_combinations(xlsx_path):
    """
    Citeste medici.xlsx si returneaza lista de tuple (cabinet, nume).
    Asteapta randul 1 = headere (cabinet, nume), restul = date.
    Doar randurile cu cabinet ne-gol sunt incluse (medicii fara cabinet
    primesc doar diplome, nu planse).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
    try:
        cab_i = headers.index('cabinet')
        nm_i  = headers.index('nume')
    except ValueError:
        raise ValueError(f'{xlsx_path}: lipsesc coloanele "cabinet" si/sau "nume" pe randul 1')

    combos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cabinet = (str(row[cab_i]).strip() if row[cab_i] else '')
        nume    = (str(row[nm_i]).strip()  if row[nm_i]  else '')
        if cabinet and nume:
            combos.append((cabinet, nume))
    return combos


def set_slide_text(slide_element, doctor, cabinet, doc_pt, cab_pt):
    """
    Inlocuieste placeholderele {{medic}} si {{cabinet}} si seteaza fonturile.
    Template-ul are 2 run-uri + 1 line break intre ele.
    sz in OOXML = sutimi de punct (80pt -> "8000").
    """
    para = slide_element.find('.//a:p', NSMAP)
    if para is None:
        return
    doc_sz = str(int(doc_pt * 100))
    cab_sz = str(int(cab_pt * 100))

    for r in para.findall('a:r', NSMAP):
        t_el = r.find('a:t', NSMAP)
        rpr  = r.find('a:rPr', NSMAP)
        if t_el is None:
            continue
        if t_el.text == PLACEHOLDER_MEDIC:
            t_el.text = doctor
            if rpr is not None:
                rpr.set('sz', doc_sz)
        elif t_el.text == PLACEHOLDER_CABINET:
            t_el.text = cabinet
            if rpr is not None:
                rpr.set('sz', cab_sz)

    # Line break: aceeasi dimensiune ca medicul (spatiu vertical intre randuri)
    for br in para.findall('a:br', NSMAP):
        rpr = br.find('a:rPr', NSMAP)
        if rpr is not None:
            rpr.set('sz', doc_sz)


def main():
    combos = load_combinations(INPUT_XLSX)
    print(f'Cabinete x medici de generat: {len(combos)}')
    print(f'Font Impact: {IMPACT_PATH or "NE-GASIT (folosesc aproximare)"}')
    print()

    prs = Presentation(str(TEMPLATE))
    template_slide = prs.slides[0]

    # Pasul 1: clonez toate slide-urile (cu placeholderele intacte).
    slides = [template_slide]
    for _ in combos[1:]:
        slides.append(clone_slide(prs, template_slide))

    # Pasul 2: aplic text + auto-fit per slide.
    for idx, (slide, (cab, doc)) in enumerate(zip(slides, combos), start=1):
        doc_pt = auto_font_pt(doc, DOC_MAX_PT, FONT_MIN_PT, SAFE_BOX_PT, IMPACT_PATH, IMPACT_RATIO)
        cab_pt = auto_font_pt(cab, CAB_MAX_PT, FONT_MIN_PT, SAFE_BOX_PT, IMPACT_PATH, IMPACT_RATIO)
        set_slide_text(slide._element, doc, cab, doc_pt, cab_pt)
        fit_note = ''
        if doc_pt < DOC_MAX_PT or cab_pt < CAB_MAX_PT:
            fit_note = ' [auto-fit]'
        print(f'  Slide {idx:3d}: [{cab}] -> [{doc}]  ({doc_pt}pt / {cab_pt}pt){fit_note}')

    OUTPUT_PPTX.parent.mkdir(exist_ok=True)
    prs.save(str(OUTPUT_PPTX))
    print(f'\nSalvat: {OUTPUT_PPTX}  ({len(combos)} slide-uri)')


if __name__ == '__main__':
    main()
